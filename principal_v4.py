# principal_v3.py
# Versión 5 - Dashboard interactivo con detalle por categoría
# Integrado en un solo archivo según el proyecto del usuario.

from ldap3 import Server, Connection, ALL, SUBTREE
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

# ==============================
# Método de envío alternativo: Outlook (perfil local)
# ==============================
def seleccionar_metodo_envio(parent):
    """Diálogo simple para elegir método de envío y, si es Outlook,
    permitir opcionalmente "Enviar como" (por ejemplo soporte@capual.cl).
    Devuelve (method, enviar_como) donde method es 'outlook' o 'smtp'.
    """
    # Detectar si Outlook (pywin32) parece disponible
    outlook_disponible = False
    try:
        import importlib
        importlib.import_module('win32com.client')  # type: ignore
        outlook_disponible = True
    except Exception:
        outlook_disponible = False

    dlg = tk.Toplevel(parent)
    setup_style(dlg)
    dlg.title("Método de envío de correos")
    dlg.resizable(False, False)
    dlg.transient(parent)
    dlg.grab_set()
    centrar_ventana(dlg, 520, 220)

    frm = ttk.Frame(dlg, padding=12)
    frm.pack(fill="both", expand=True)

    ttk.Label(frm, text="Elige cómo enviar los correos:").pack(anchor="w")
    metodo_var = tk.StringVar(value="outlook" if outlook_disponible else "smtp")
    rb1 = ttk.Radiobutton(frm, text="Usar Outlook (más simple)", value="outlook", variable=metodo_var)
    rb2 = ttk.Radiobutton(frm, text="Usar SMTP (pedirá usuario/clave)", value="smtp", variable=metodo_var)
    rb1.pack(anchor="w", pady=(6,2))
    rb2.pack(anchor="w", pady=(0,10))

    if not outlook_disponible:
        try:
            rb1.state(["disabled"])  # deshabilitar si no disponible
        except Exception:
            pass
        ttk.Label(frm, text="Outlook no disponible en este equipo (falta pywin32 o Outlook).",
                  foreground="#a33").pack(anchor="w")

    # Campo opcional Enviar como (solo para Outlook)
    enviar_como_var = tk.StringVar(value="")
    enviar_como_row = ttk.Frame(frm)
    enviar_como_row.pack(fill="x", pady=(10,0))
    ttk.Label(enviar_como_row, text="Enviar como (opcional):").pack(side="left")
    e_enviar_como = ttk.Entry(enviar_como_row, textvariable=enviar_como_var, width=34)
    e_enviar_como.pack(side="left", padx=(8,0))

    ayuda = ttk.Label(frm, text="Deja vacío para usar tu propia cuenta de Outlook.\n"
                             "Para usar soporte@capual.cl necesitas permiso de 'Enviar como' o se enviará 'en nombre de'.",
                      font=("Segoe UI", 9))
    ayuda.pack(anchor="w", pady=(6,0))

    def toggle_enviar_como(*args):
        if metodo_var.get() == "outlook":
            try:
                e_enviar_como.state(["!disabled"])  # habilitar
            except Exception:
                pass
        else:
            enviar_como_var.set("")
            try:
                e_enviar_como.state(["disabled"])  # deshabilitar
            except Exception:
                pass

    metodo_var.trace_add("write", lambda *a: toggle_enviar_como())
    toggle_enviar_como()

    btns = ttk.Frame(frm)
    btns.pack(fill="x", pady=(12,0))
    result = {"ok": False}

    def aceptar():
        result["ok"] = True
        dlg.destroy()

    def cancelar():
        result["ok"] = False
        dlg.destroy()

    ttk.Button(btns, text="Cancelar", command=cancelar).pack(side="right", padx=(6,0))
    ttk.Button(btns, text="Continuar", command=aceptar).pack(side="right")

    dlg.bind("<Escape>", lambda e: cancelar())
    dlg.wait_window()

    if not result["ok"]:
        return None, None
    return metodo_var.get(), (enviar_como_var.get().strip() or None)


def _get_outlook_app():
    """Intenta inicializar Outlook COM con varios ProgID y EnsureDispatch.
    Lanza la última excepción si no es posible.
    """
    try:
        import pythoncom  # type: ignore
        pythoncom.CoInitialize()
    except Exception:
        # Continuar incluso si falla; win32com suele inicializar por nosotros
        pass

    try:
        import win32com.client  # type: ignore
        from win32com.client import gencache  # type: ignore
    except Exception as e:
        raise e

    progids = [
        "Outlook.Application",
        "Outlook.Application.16",
        "Outlook.Application.15",
    ]
    last_err = None
    for pid in progids:
        try:
            # EnsureDispatch regenera wrappers si el cache está corrupto
            return gencache.EnsureDispatch(pid)
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise RuntimeError("No se pudo localizar Outlook (COM no registrado)")


def enviar_correos_via_outlook(usuarios, parent, enviar_como=None):
    """Envía correos usando Outlook (perfil local). Soporta 'Enviar como' si hay permisos.
    Devuelve True si envió al menos uno.
    """
    # Verificar selección
    # Cargar COM de Outlook
    try:
        outlook = _get_outlook_app()
    except Exception as e:
        # Mensaje claro para causas típicas de REGDB_E_CLASSNOTREG
        msg = (
            "No se pudo inicializar Outlook (COM).\n\n"
            f"Detalle: {e}\n\n"
            "Posibles causas y soluciones:\n"
            "• Outlook no está instalado o no se abrió al menos una vez.\n"
            "• La edición de Outlook de Microsoft Store no expone COM. Instala Microsoft 365 Apps (Click-to-Run).\n"
            "• Reparar Office desde Configuración > Aplicaciones > Microsoft 365 > Modificar.\n"
            "• Asegúrate de tener pywin32 instalado en el entorno de ejecución.\n"
        )
        messagebox.showerror("Outlook no disponible", msg, parent=parent)
        return False

    if not usuarios:
        messagebox.showwarning("Aviso", "No hay usuarios seleccionados para enviar correos.", parent=parent)
        return False

    if not messagebox.askyesno("Confirmar envío",
                               f"¿Desea enviar correos a {len(usuarios)} usuarios?\n\nVía: Outlook" +
                               (f"\nEnviar como: {enviar_como}" if enviar_como else ""), parent=parent):
        return False

    progress_win = tk.Toplevel(parent)
    setup_style(progress_win)
    progress_win.title("Enviando correos… (Outlook)")
    progress_win.transient(parent)
    progress_win.grab_set()
    centrar_ventana(progress_win, 460, 150)

    frm = ttk.Frame(progress_win, padding=12)
    frm.pack(fill="both", expand=True)
    lbl = ttk.Label(frm, text="Preparando…")
    lbl.pack(fill="x", pady=(0,8))
    pbar = ttk.Progressbar(frm, mode="determinate", maximum=len(usuarios))
    pbar.pack(fill="x")
    cancel = tk.BooleanVar(value=False)
    ttk.Button(frm, text="Cancelar", command=lambda: cancel.set(True)).pack(pady=(10,0))

    enviados = 0
    try:
        # ns = outlook.GetNamespace("MAPI")  # si se requiere: ns.Logon(None, None, True, False)

        for idx, u in enumerate(usuarios, start=1):
            if cancel.get():
                break
            correo = u.get("correo")
            if not correo or "@" not in correo:
                continue

            try:
                mail = outlook.CreateItem(0)  # olMailItem
                mail.To = correo
                mail.Subject = "⚠️ Aviso: Tu contraseña está próxima a expirar"
                if enviar_como:
                    # Enviar como / en nombre de (requiere permisos en Exchange)
                    try:
                        mail.SentOnBehalfOfName = enviar_como
                    except Exception:
                        pass

                html_body = f"""
                <html>
                <body style="font-family:Segoe UI, sans-serif; color:#333;">
                    <p>Estimado/a <b>{u.get('nombre','')}</b>,</p>
                    <p>Tu contraseña expira en <b>{u.get('dias','-')} días</b> (el {u.get('expira','-')}).<br>
                    Por favor, actualízala antes de que caduque para evitar bloqueos de acceso.</p>
                    <p><b>Para cambiar tu contraseña:</b><br>
                    Presiona <i>Ctrl + Alt + Supr</i> y selecciona la opción "Cambiar contraseña".</p>
                    <p style="text-align:center;">
                        <img src="cid:img_teclas" alt="Instrucciones Ctrl+Alt+Supr" width="420">
                    </p>
                    <p>Si tienes problemas, comunícate con:<br>
                    - Eduardo L. (Nexo 4006)<br>
                    - Ignacio C. (Nexo 4018)<br>
                    Departamento de Servicios TI</p>
                    <p>🏢 Departamento: {u.get('departamento','No especificado')}<br>
                    👤 Usuario: {u.get('usuario','')}</p>
                    <p>Saludos cordiales,<br>
                    <b>Departamento de Soporte TI</b><br>
                    Capual - Cooperativa de Ahorro y Crédito</p>
                </body>
                </html>
                """

                # Adjuntar imagen inline con Content-ID
                if os.path.exists(IMG_PATH):
                    attach = mail.Attachments.Add(IMG_PATH)
                    try:
                        # PR_ATTACH_CONTENT_ID = 0x3712001F
                        pa = attach.PropertyAccessor
                        pa.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "img_teclas")
                    except Exception:
                        pass

                mail.HTMLBody = html_body
                mail.Send()
                enviados += 1
            except Exception:
                # Continuar con el siguiente destinatario
                pass

            lbl.config(text=f"Enviando {idx}/{len(usuarios)}…")
            pbar['value'] = idx
            progress_win.update_idletasks()

    except Exception as e:
        # -2147221005 => REGDB_E_CLASSNOTREG (Clase no registrada)
        h = getattr(e, 'hresult', None)
        if h == -2147221005 or "clase" in str(e).lower():
            sugerencia = (
                "Clase COM no registrada para Outlook.\n\n"
                "Sugerencias:\n"
                "• Evita la versión de Microsoft Store de Office; usa Microsoft 365 Apps (Click-to-Run).\n"
                "• Ejecuta Outlook una vez para que se registre.\n"
                "• Repara Office.\n"
            )
        else:
            sugerencia = ""
        messagebox.showerror(
            "Outlook",
            f"No se pudo enviar con Outlook:\n{e}\n\n{sugerencia}Prueba con SMTP.",
            parent=progress_win,
        )
    finally:
        try:
            progress_win.destroy()
        except Exception:
            pass

    if enviados > 0:
        messagebox.showinfo("Envío finalizado", f"Correos enviados: {enviados}", parent=parent)
        return True
    return False
import tkinter as tk
from tkinter import messagebox, ttk, simpledialog, filedialog
import sys
import matplotlib
# usar backend TkAgg para interacción en la UI
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os, sys, re

# ==============================
# CONFIGURACIÓN (ajusta según tu entorno)
# ==============================
AD_SERVER = 'ldaps://SRV_DC01_NEW.capual.cl'
BASE_DN = 'DC=capual,DC=cl'

# Limitar la búsqueda solo a estas OUs bajo OU=Capual
ALLOWED_OUS = [
    "OU=Areas de Apoyo,OU=Capual,DC=capual,DC=cl",
    "OU=Directorio,OU=Capual,DC=capual,DC=cl",
    "OU=Gerencia,OU=Capual,DC=capual,DC=cl",
    "OU=Oficinas,OU=Capual,DC=capual,DC=cl",
]


# Las credenciales de envío se solicitarán en tiempo de ejecución (no se almacenan en el código)
SMTP_REMITENTE = "printservice@capual.cl"
SMTP_PASSWORD = "PSD34$/srvc123."
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587

# Cache opcional de credenciales solo para la sesión actual (si el usuario lo permite)
_SMTP_CACHE = {"remitente": None, "password": None, "from_visible": None}

APP_CREDITOS = """App creada por Eduardo 'PaladynamoX' Lizama C.
Versión 5.0.0 - Año 2025"""

# Detecta si se está ejecutando desde .exe o desde .py
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

IMG_PATH = os.path.join(BASE_DIR, "img_teclas.png")
LOGO_PATH = os.path.join(BASE_DIR, "logo_capual_antiguo.png")
FAREWELL_LOGO_PATH = os.path.join(BASE_DIR, "kuriboh_logo_despedida.png")


# ==============================
# UTILIDADES UI
# ==============================
def centrar_ventana(ventana, ancho, alto):
    ventana.update_idletasks()
    x = (ventana.winfo_screenwidth() // 2) - (ancho // 2)
    y = (ventana.winfo_screenheight() // 2) - (alto // 2)
    ventana.geometry(f"{ancho}x{alto}+{x}+{y}")


# Estilo visual unificado (ligero) para toda la app
def setup_style(root):
    """Aplica un estilo visual con tonos verde metalizado y corrige
    el mapeo de colores para evitar textos invisibles en botones."""
    try:
        style = ttk.Style(root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        # Paleta (verde claro metalizado)
        base_bg = "#eef6f1"   # fondo muy claro verdoso
        border = "#b6d6c3"

        # Fondo general
        style.configure("TFrame", background=base_bg)
        style.configure("TLabelframe", background=base_bg)
        style.configure("TLabel", background=base_bg, foreground="#20302a", font=("Segoe UI", 10))

        # Botones: volvemos a un estilo seguro (sin colorear fondo) para que SIEMPRE se
        # renderice el texto correctamente en Windows.
        style.configure("TButton", font=("Segoe UI", 10), padding=(10, 6), foreground="#222")
        style.map(
            "TButton",
            background=[("active", "#e6f3ec")],  # leve tinte verde en hover
            foreground=[
                ("disabled", "#777"),
                ("active", "#222"),
                ("!disabled", "#222"),
            ],
        )

        # Tabla
        style.configure(
            "Treeview",
            font=("Segoe UI", 10),
            rowheight=24,
            background="#ffffff",
            fieldbackground="#ffffff",
            bordercolor=border,
        )
        style.configure(
            "Treeview.Heading",
            font=("Segoe UI Semibold", 10),
            background="#d8efe2",
            foreground="#1f2a26",
            bordercolor=border,
        )
    except Exception:
        pass


# Corrección simple de textos con tildes mal codificadas (mojibake)
def fix_text_encoding(s: str) -> str:
    if not isinstance(s, str):
        return s
    # Heurística: patrones comunes de mojibake UTF-8 interpretado como Latin-1
    if any(p in s for p in ("Ã", "Â", "Ð", "Þ")):
        try:
            return s.encode("latin-1", errors="ignore").decode("utf-8", errors="ignore")
        except Exception:
            return s
    return s


# Ajusta la altura de la tabla y la ventana según el número de registros
def auto_ajustar_altura(*args, **kwargs):
    # Desactivado por solicitud: mantenemos firma por compatibilidad
    return


# ==============================
# UTILIDAD: Comportamiento unificado de tablas (Treeview)
# ==============================
def _parse_date_ddmmyyyy(s: str):
    try:
        return datetime.strptime(s.strip(), "%d/%m/%Y %H:%M")
    except Exception:
        try:
            return datetime.strptime(s.strip(), "%d/%m/%Y")
        except Exception:
            return None


def make_treeview_standard(tree: ttk.Treeview, cols: tuple, item_to_user: dict, on_view_properties, seleccion: dict):
    """Aplica a un Treeview:
    - Orden asc/desc al click en encabezados
    - Selección por checkbox en columna 'Sel'
    - Encabezado 'Sel' alterna seleccionar/deseleccionar todo
    - Resalte de fila activa al click
    - Doble clic abre propiedades (on_view_properties)
    """
    # Estilo de fila activa
    try:
        tree.tag_configure("row_active", background="#cce5ff")
    except Exception:
        pass

    sort_state = {c: False for c in cols}  # False: asc, True: desc
    select_all_state = {"all": False}
    active_iid = {"iid": None}

    def _value_for_sort(iid, col):
        val = tree.set(iid, col)
        col_lower = str(col).lower()
        # Días -> int
        if "día" in col_lower:
            try:
                return int(str(val).strip())
            except Exception:
                return float("inf")
        # Fecha -> datetime
        if "fecha" in col_lower:
            dt = _parse_date_ddmmyyyy(str(val))
            return dt or datetime.max
        # default: string casefold
        return str(val).casefold()

    def ordenar(col):
        reverse = sort_state[col]
        data = [( _value_for_sort(iid, col), iid) for iid in tree.get_children("")]
        data.sort(key=lambda t: t[0], reverse=reverse)
        for idx, (_v, iid) in enumerate(data):
            tree.move(iid, "", idx)
        sort_state[col] = not reverse

    def toggle_all_sel():
        # Solo aplica si existe columna Sel
        if not cols or str(cols[0]).lower().startswith("sel") is False:
            return
        new_state = not select_all_state["all"]
        for iid in tree.get_children(""):
            seleccion[iid] = new_state
            tree.set(iid, cols[0], "☑" if new_state else "☐")
        select_all_state["all"] = new_state

    # Encabezados
    for c in cols:
        if str(c).lower().startswith("sel"):
            tree.heading(c, text="Sel", command=toggle_all_sel)
        else:
            tree.heading(c, command=lambda c=c: ordenar(c))

    # Click: toggle checkbox o marcar fila activa
    def on_click(event):
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        iid = tree.identify_row(event.y)
        col_id = tree.identify_column(event.x)  # '#1', '#2', ...
        if not iid:
            return
        # Limpiar activo
        if active_iid["iid"] and tree.exists(active_iid["iid"]):
            tree.item(active_iid["iid"], tags=())
        # Si clic en primera columna (Sel)
        if col_id == "#1" and str(cols[0]).lower().startswith("sel"):
            seleccion[iid] = not seleccion.get(iid, False)
            tree.set(iid, cols[0], "☑" if seleccion[iid] else "☐")
        else:
            # Resalte fila activa
            tree.item(iid, tags=("row_active",))
            active_iid["iid"] = iid

    tree.bind("<Button-1>", on_click)

    # Doble clic: ver propiedades
    def on_dclick(event):
        iid = tree.identify_row(event.y)
        if not iid:
            return
        user = item_to_user.get(iid)
        if user is not None:
            try:
                on_view_properties(user)
            except Exception:
                pass
    tree.bind("<Double-1>", on_dclick)

# ==============================
# UTILIDAD: Exportar Treeview a Excel con estilo, logo y resumen
# ==============================
def _ensure_package_silent(pkg: str, import_name: str = None) -> bool:
    """Intenta importar un paquete y si falla, lo instala en silencio con pip.
    Devuelve True si el paquete está disponible al finalizar."""
    import importlib, subprocess, sys
    name = import_name or pkg
    try:
        importlib.import_module(name)
        return True
    except Exception:
        pass

    # Evitar intentos de instalación en ejecutables congelados (.exe)
    if getattr(sys, 'frozen', False):
        return False

    cmds = [
        [sys.executable, "-m", "pip", "install", pkg, "--quiet", "--disable-pip-version-check"],
        [sys.executable, "-m", "pip", "install", pkg, "--user", "--quiet", "--disable-pip-version-check"],
    ]
    for cmd in cmds:
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            import importlib
            importlib.invalidate_caches()
            importlib.import_module(name)
            return True
        except Exception:
            continue
    return False


def _ensure_excel_deps(parent_window=None) -> bool:
    """Asegura 'openpyxl' y 'Pillow' (para insertar imágenes)."""
    ok_openpyxl = _ensure_package_silent("openpyxl")
    ok_pillow = _ensure_package_silent("Pillow", import_name="PIL")
    if not ok_openpyxl:
        try:
            messagebox.showerror(
                "Dependencia faltante",
                "No fue posible cargar o instalar 'openpyxl'.\n\n" \
                "Si estás ejecutando un .exe, incluye 'openpyxl' en el empaquetado."
            )
        except Exception:
            pass
        return False
    # Pillow es opcional (solo para logo). Si falla, seguimos sin imagen.
    return True


def export_tree_to_excel(parent_window, tree: ttk.Treeview, titulo: str = "Reporte"):
    # Intentar importar dependencias; si faltan, instalarlas silenciosamente cuando sea posible.
    if not _ensure_excel_deps(parent_window):
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        XLImage = None

    # Elegir ruta
    path = filedialog.asksaveasfilename(
        parent=parent_window,
        title="Guardar como",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"{titulo.replace(' ', '_')}.xlsx",
    )
    if not path:
        return

    # Preparar datos desde el Treeview
    cols = list(tree["columns"]) or []
    include_idx = [i for i, c in enumerate(cols) if str(c).strip().lower() != "sel"]
    headers = [cols[i] for i in include_idx]
    rows = [tree.item(iid, 'values') for iid in tree.get_children()]
    rows = [[r[i] for i in include_idx] for r in rows]

    # Calcular resumen por categorías a partir de "Días restantes"
    def _parse_int_safe(v):
        try:
            return int(str(v).strip())
        except Exception:
            return None

    dias_col_idx = None
    for idx, h in enumerate(headers):
        if "día" in str(h).lower():
            dias_col_idx = idx
            break
    counts = {"Bien (16-90)": 0, "Próximos (1-15)": 0, "Expirados (<=0)": 0, "Sin dato": 0}
    if dias_col_idx is not None:
        for r in rows:
            v = _parse_int_safe(r[dias_col_idx])
            if v is None:
                counts["Sin dato"] += 1
            elif v <= 0:
                counts["Expirados (<=0)"] += 1
            elif 1 <= v <= 15:
                counts["Próximos (1-15)"] += 1
            elif 16 <= v <= 90:
                counts["Bien (16-90)"] += 1
            else:
                # Fuera de rango superior, no contamos o podríamos sumarlo en Bien
                counts["Bien (16-90)"] += 0
    else:
        # No hay columna de días; solo creamos hoja de datos
        pass

    wb = openpyxl.Workbook()
    # Hoja de resumen primero
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws = wb.create_sheet("Datos")

    # Estilos base
    verde = "27ae60"  # sin # para fills
    verde_claro = "d8efe2"
    gris_claro = "f7fbf9"
    thin = Side(style="thin", color="b6d6c3")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Logo (si es posible) para ambas hojas, con reescalado nítido
    top_row = 1
    start_col_for_title = 3  # dejamos espacio para el logo en la columna A/B
    try:
        from io import BytesIO
        try:
            from PIL import Image as PILImage  # type: ignore
        except Exception:
            PILImage = None  # Pillow no disponible
        try:
            from openpyxl.drawing.image import Image as XLImage  # type: ignore
        except Exception:
            XLImage = None

        def _add_logo(ws_local, cell="A1", max_w=220, max_h=70):
            if not os.path.exists(LOGO_PATH) or XLImage is None:
                return
            if PILImage is None:
                # Sin Pillow: agregar sin reescalar (mejor que nada)
                try:
                    ws_local.add_image(XLImage(LOGO_PATH))
                except Exception:
                    pass
                return
            try:
                im = PILImage.open(LOGO_PATH)
                im = im.convert("RGBA")
                resample = getattr(PILImage, "LANCZOS", getattr(PILImage, "BICUBIC", 1))
                im.thumbnail((max_w, max_h), resample)
                bio = BytesIO()
                im.save(bio, format="PNG")
                bio.seek(0)
                img = XLImage(bio)
                img.anchor = cell
                ws_local.add_image(img)
            except Exception:
                pass

        _add_logo(ws_res, "A1", 220, 70)
        _add_logo(ws, "A1", 220, 70)
    except Exception:
        pass

    last_col = max(len(headers), 1)
    # Título (fila 1)
    ws.merge_cells(start_row=top_row, start_column=start_col_for_title, end_row=top_row, end_column=last_col+1)
    c = ws.cell(row=top_row, column=start_col_for_title, value=titulo)
    c.font = Font(size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=verde)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Subtítulo (fecha) fila 2
    ws.merge_cells(start_row=top_row+1, start_column=start_col_for_title, end_row=top_row+1, end_column=last_col+1)
    c2 = ws.cell(row=top_row+1, column=start_col_for_title, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c2.font = Font(size=10, color="24543b")
    c2.alignment = Alignment(horizontal="center")

    # Encabezados (fila 4)
    header_row = top_row + 3
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = Font(bold=True, color="1f2a26")
        cell.fill = PatternFill("solid", fgColor=verde_claro)
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde

    # Datos
    data_start = header_row + 1
    for i, r in enumerate(rows, start=data_start):
        for j, v in enumerate(r, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = borde
            # Alinear ciertas columnas
            header = headers[j-1].lower()
            if "día" in header:
                cell.alignment = Alignment(horizontal="center")
            elif "fecha" in header:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
        # Zebra stripe
        if (i - data_start) % 2 == 0:
            for j in range(1, len(headers)+1):
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor=gris_claro)

    # Auto ancho de columnas
    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        max_len = len(str(h))
        for i in range(data_start, data_start + len(rows)):
            val = ws.cell(row=i, column=j).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)

    # Filtros, panes congelados
    ws.auto_filter.ref = f"A{header_row}:" + get_column_letter(len(headers)) + f"{header_row}"
    ws.freeze_panes = f"A{data_start}"

    # Formato condicional para "Días restantes" (<=0 rojo, 1-15 amarillo, 16-90 verde)
    try:
        dias_idx = next((idx+1 for idx, h in enumerate(headers) if "día" in str(h).lower()), None)
        if dias_idx and rows:
            col_letter = get_column_letter(dias_idx)
            rng = f"{col_letter}{data_start}:{col_letter}{data_start + len(rows) - 1}"
            rojo = PatternFill("solid", fgColor="f8d7da")
            amarillo = PatternFill("solid", fgColor="fff3cd")
            verde_fill = PatternFill("solid", fgColor="d4edda")

            ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=['0'], fill=rojo))
            ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['1','15'], fill=amarillo))
            ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['16','90'], fill=verde_fill))

            # Número entero para la columna de días
            for i in range(data_start, data_start + len(rows)):
                ws.cell(row=i, column=dias_idx).number_format = "0"
    except Exception:
        pass

    # Construir hoja Resumen con KPI vistosas y gráfico donut
    try:
        # Título y subtítulo
        ws_res.merge_cells(start_row=top_row, start_column=start_col_for_title, end_row=top_row, end_column=start_col_for_title+5)
        c = ws_res.cell(row=top_row, column=start_col_for_title, value=f"Resumen – {titulo}")
        c.font = Font(size=16, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=verde)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws_res.merge_cells(start_row=top_row+1, start_column=start_col_for_title, end_row=top_row+1, end_column=start_col_for_title+5)
        c2 = ws_res.cell(row=top_row+1, column=start_col_for_title, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        c2.font = Font(size=10, color="24543b")
        c2.alignment = Alignment(horizontal="center")

        # KPIs
        total_reg = sum(counts.values()) if counts else len(rows)
        def _pct(n):
            try:
                return f"{(n/total_reg*100):.1f}%" if total_reg else "0%"
            except Exception:
                return "-"

        kpi_defs = [
            ("Usuarios totales", total_reg, "24543b", "e9f4ef"),
            ("Bien (16-90)", counts.get("Bien (16-90)", 0), "1e8449", "d8efe2"),
            ("Próximos (1-15)", counts.get("Próximos (1-15)", 0), "ba8b00", "fff4cc"),
            ("Expirados (<=0)", counts.get("Expirados (<=0)", 0), "a93226", "f9d6d5"),
        ]
        start_col_cards = 2
        row_cards = top_row + 3
        span_w, span_h = 3, 3
        for i, (label, val, color_txt, color_bg) in enumerate(kpi_defs):
            c1 = start_col_cards + i * (span_w + 1)
            r1 = row_cards
            ws_res.merge_cells(start_row=r1, start_column=c1, end_row=r1+span_h-1, end_column=c1+span_w-1)
            cell = ws_res.cell(row=r1, column=c1)
            cell.value = f"{label}\n{val} ({_pct(val)})"
            cell.font = Font(size=12, bold=True, color=color_txt)
            cell.fill = PatternFill("solid", fgColor=color_bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for rr in range(r1, r1+span_h):
                for cc in range(c1, c1+span_w):
                    ws_res.cell(row=rr, column=cc).border = borde

        # Datos para gráfico donut + tabla de estados con porcentaje
        dist_row = row_cards + span_h + 2
        ws_res.cell(row=dist_row, column=2, value="Estado").font = Font(bold=True)
        ws_res.cell(row=dist_row, column=3, value="Cantidad").font = Font(bold=True)
        ws_res.cell(row=dist_row, column=4, value="Porcentaje").font = Font(bold=True)
        estados = ["Bien (16-90)", "Próximos (1-15)", "Expirados (<=0)"]
        for i, est in enumerate(estados, start=1):
            ws_res.cell(row=dist_row+i, column=2, value=est)
            val = counts.get(est, 0)
            ws_res.cell(row=dist_row+i, column=3, value=val).number_format = "0"
            # porcentaje como valor numérico (0..1)
            pct = (val / total_reg) if total_reg else 0
            c_pct = ws_res.cell(row=dist_row+i, column=4, value=pct)
            c_pct.number_format = "0.0%"

        # Barras de datos visuales en la columna de porcentaje
        try:
            from openpyxl.formatting.rule import DataBarRule
            rng_pct = f"D{dist_row+1}:D{dist_row+len(estados)}"
            rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                               color="99CC00", showValue=True)
            ws_res.conditional_formatting.add(rng_pct, rule)
        except Exception:
            pass

        try:
            from openpyxl.chart import DoughnutChart, Reference
            donut = DoughnutChart()
            donut.title = "Distribución por estado"
            donut.holeSize = 60
            data = Reference(ws_res, min_col=3, min_row=dist_row, max_row=dist_row+len(estados))
            labels = Reference(ws_res, min_col=2, min_row=dist_row+1, max_row=dist_row+len(estados))
            donut.add_data(data, titles_from_data=True)
            donut.set_categories(labels)
            ws_res.add_chart(donut, f"E{dist_row}")
        except Exception:
            pass

        # Link de navegación a hoja Datos
        link_row = dist_row + len(estados) + 1
        go = ws_res.cell(row=link_row, column=2, value="Ir a hoja 'Datos' →")
        go.hyperlink = "#'Datos'!A1"
        go.font = Font(color="0563C1", underline="single", bold=True)

        # Top 5 departamentos (<=15 días) si existen columnas
        dept_top_row = dist_row + len(estados) + 3

        try:
            dept_idx = None
            for idx, h in enumerate(headers):
                if str(h).strip().lower() == "departamento":
                    dept_idx = idx
                    break
            if dept_idx is not None and dias_col_idx is not None:
                tmp = {}
                for r in rows:
                    try:
                        dias = int(str(r[dias_col_idx]).strip())
                    except Exception:
                        continue
                    if dias <= 15:
                        d = str(r[dept_idx] or "(Sin departamento)")
                        tmp[d] = tmp.get(d, 0) + 1
                top5 = sorted(tmp.items(), key=lambda x: x[1], reverse=True)[:5]
            else:
                top5 = []

            ws_res.cell(row=dept_top_row, column=2, value="Top departamentos (<=15 días)").font = Font(bold=True)
            ws_res.cell(row=dept_top_row+1, column=2, value="Departamento").font = Font(bold=True)
            ws_res.cell(row=dept_top_row+1, column=3, value="Cantidad").font = Font(bold=True)
            for i, (dname, n) in enumerate(top5, start=1):
                ws_res.cell(row=dept_top_row+1+i, column=2, value=dname)
                ws_res.cell(row=dept_top_row+1+i, column=3, value=n).number_format = "0"
            for rr in range(dept_top_row+1, dept_top_row+2+max(len(top5),1)):
                for cc in (2,3):
                    ws_res.cell(row=rr, column=cc).border = borde
                    if rr == dept_top_row+1:
                        ws_res.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=gris_claro)
        except Exception:
            pass

        # Top 10 más urgentes (incluye expirados y próximos)
        urgent_row = dept_top_row + 2 + max(len(top5), 1) + 2

        try:
            # Indices de columnas en Datos
            idx_user = next((i for i,h in enumerate(headers) if str(h).lower().startswith("usuario")), None)
            idx_name = next((i for i,h in enumerate(headers) if str(h).lower().startswith("nombre")), None)
            idx_dept = next((i for i,h in enumerate(headers) if "depart" in str(h).lower()), None)
            idx_fecha = next((i for i,h in enumerate(headers) if "fecha" in str(h).lower()), None)
            urg = []
            if dias_col_idx is not None:
                for r in rows:
                    try:
                        d = int(str(r[dias_col_idx]).strip())
                    except Exception:
                        continue
                    if d <= 15:  # próximos o expirados
                        urg.append((d,
                                    (r[idx_user] if idx_user is not None else ""),
                                    (r[idx_name] if idx_name is not None else ""),
                                    (r[idx_dept] if idx_dept is not None else ""),
                                    (r[idx_fecha] if idx_fecha is not None else "")))
                urg.sort(key=lambda x: x[0])
                urg = urg[:10]
            # Título
            ws_res.cell(row=urgent_row, column=2, value="Top 10 más urgentes (≤15 días)").font = Font(bold=True)
            # Encabezados
            hdrs = ["Días", "Usuario", "Nombre", "Departamento", "Fecha"]
            for j, htxt in enumerate(hdrs, start=2):
                cellh = ws_res.cell(row=urgent_row+1, column=j, value=htxt)
                cellh.font = Font(bold=True)
                cellh.fill = PatternFill("solid", fgColor=verde_claro)
                cellh.border = borde
            # Filas
            for i, item in enumerate(urg, start=1):
                d, u, n, dp, f = item
                vals = [d, u, n, dp, f]
                for j, v in enumerate(vals, start=2):
                    cellv = ws_res.cell(row=urgent_row+1+i, column=j, value=v)
                    cellv.border = borde
                    if j == 2:
                        cellv.number_format = "0"
            # Zebra
            total_urg = max(len(urg), 1)
            for r in range(urgent_row+2, urgent_row+2+total_urg):
                if (r - (urgent_row+2)) % 2 == 0:
                    for j in range(2, 7):
                        ws_res.cell(row=r, column=j).fill = PatternFill("solid", fgColor=gris_claro)
        except Exception:
            pass

        # Leyenda y definiciones
        info_row = urgent_row + 3 + max(len(urg) if 'urg' in locals() else 0, 1)
        ws_res.merge_cells(start_row=info_row, start_column=2, end_row=info_row+2, end_column=8)
        note = ws_res.cell(row=info_row, column=2, value=(
            "Definiciones: Bien=16–90 días, Próximos=1–15 días, Expirados=≤0 días. "
            "Este informe resume la salud de contraseñas. Use el enlace 'Ir a hoja Datos' para filtrar y revisar casos."
        ))
        note.alignment = Alignment(wrap_text=True, vertical="top")
        note.fill = PatternFill("solid", fgColor=gris_claro)
        note.border = borde

        # Ajustes de columnas en Resumen
        for j in range(2, 10):
            try:
                ws_res.column_dimensions[get_column_letter(j)].width = 20
            except Exception:
                pass
    except Exception:
        pass

    # Guardar al final, ya con Resumen
    try:
        wb.save(path)
        messagebox.showinfo("Exportar a Excel", "Archivo Excel exportado correctamente.", parent=parent_window)
    except Exception as e:
        messagebox.showerror("Exportar a Excel", f"No se pudo guardar el archivo:\n{e}", parent=parent_window)


def despedida_final(segundos: int = 5):
    """Muestra una ventana de despedida centrada, sin botones ni contador, con logo, y cierra al cabo de 'segundos'."""
    try:
        # Ocultar ventana raíz si existe para que solo se vea la despedida
        if root_all and root_all.winfo_exists():
            try:
                root_all.withdraw()
            except Exception:
                pass

        # Crear diálogo de despedida
        dlg = tk.Toplevel()
        setup_style(dlg)
        dlg.title("Despedida")
        dlg.resizable(False, False)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)  # impedir cierre manual
        width, height = 520, 260
        dlg.transient(None)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=16)
        frm.pack(fill="both", expand=True)

        # Mensaje de créditos/firma
        msg = ttk.Label(
            frm,
            text=f"Gracias por usar esta aplicación.\n{APP_CREDITOS}",
            anchor="center",
            justify="center"
        )
        msg.pack(fill="x", pady=(0, 10))

        # Logo personal (escalado para caber en la ventana)
        logo_label = ttk.Label(frm)
        logo_label.pack(pady=(0, 4))
        try:
            from PIL import Image, ImageTk  # type: ignore
            if os.path.exists(FAREWELL_LOGO_PATH):
                with Image.open(FAREWELL_LOGO_PATH) as im:
                    max_w = width - 64  # margen dentro del cuadro
                    max_h = 110
                    im.thumbnail((max_w, max_h), Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS)
                    photo = ImageTk.PhotoImage(im)
                    logo_label.configure(image=photo)
                    # Guardar referencia para evitar GC
                    dlg._farewell_photo = photo
        except Exception:
            # Si PIL no está disponible o la imagen falla, continuar sin logo
            pass

        # Centrar después de construir contenido real
        try:
            dlg.update_idletasks()
            real_w = dlg.winfo_width() or width
            real_h = dlg.winfo_height() or height
            centrar_ventana(dlg, int(real_w), int(real_h))
            # Recentrar tras un pequeño retraso por si el sistema ajusta bordes/título
            dlg.after(120, lambda: centrar_ventana(dlg, dlg.winfo_width(), dlg.winfo_height()))
        except Exception:
            centrar_ventana(dlg, width, height)

        def finalizar():
            try:
                dlg.destroy()
            except Exception:
                pass
            try:
                if root_all and root_all.winfo_exists():
                    root_all.destroy()
            except Exception:
                pass
            sys.exit(0)

        # Programar cierre automático sin mostrar contador
        dlg.after(max(1000, int(segundos) * 1000), finalizar)

        # Asegurar que quede al frente y centrado
        try:
            dlg.lift()
            dlg.attributes('-topmost', True)
            dlg.after(200, lambda: dlg.attributes('-topmost', False))
            # Recentrar otra vez por seguridad tras el cambio de z-order/topmost
            dlg.after(220, lambda: centrar_ventana(dlg, dlg.winfo_width(), dlg.winfo_height()))
        except Exception:
            pass
    except Exception:
        # Cierre inmediato en caso de imprevistos
        try:
            if root_all and root_all.winfo_exists():
                root_all.destroy()
        except Exception:
            pass
        sys.exit(0)


def confirmar_y_cerrar(ventana):
    if messagebox.askyesno("Confirmar salida", "¿Deseas salir del programa?"):
        despedida_final(5)
    else:
        return

# ==============================
# FUNCIONES LDAP / CONSULTAS
# ==============================
def conectar_ldap(username, password):
    user = username
    if "@" not in user:
        user = f"{user}@capual.cl"
    try:
        server = Server(AD_SERVER, get_info=ALL)
        conn = Connection(server, user=user, password=password, authentication='SIMPLE', auto_bind=True)
        return conn
    except Exception as e:
        raise


def msds_to_datetime(msds_value):
    try:
        if not msds_value or int(msds_value) <= 0:
            return None
        expiry_date = datetime.fromtimestamp(int(msds_value) / 1e7 - 11644473600)
        return expiry_date
    except Exception:
        return None


def consultar_usuarios(conn):
    """
    Consulta los usuarios válidos desde el AD, restringido a las OUs indicadas.
    - Excluye usuarios sin correo.
    - Excluye usuarios con descripción: Auxiliar, Vigilante Privado o Guardia.
    - Solo busca en las subcarpetas válidas de 'Capual'.
    """

    filter_query = (
        "(&"
        "(objectCategory=person)"
        "(objectClass=user)"
        "(!(userAccountControl:1.2.840.113556.1.4.803:=2))"        # no deshabilitados
        "(!(userAccountControl:1.2.840.113556.1.4.803:=65536))"     # no 'password never expires'
        "(!(sAMAccountName=*$))"                                    # no cuentas de servicio
        "(!(sAMAccountName=Administrador))"
        "(mail=*)"                                                  # ✅ solo usuarios con correo
        ")"
    )

    attributes = [
        "sAMAccountName",
        "displayName",
        "mail",
        "msDS-UserPasswordExpiryTimeComputed",
        "department",
        "description",
        "distinguishedName",
    ]

    # 🔒 Solo en estas unidades organizativas
    ALLOWED_OUS = [
        "OU=Areas de Apoyo,OU=Capual,DC=capual,DC=cl",
        "OU=Directorio,OU=Capual,DC=capual,DC=cl",
        "OU=Gerencia,OU=Capual,DC=capual,DC=cl",
        "OU=Oficinas,OU=Capual,DC=capual,DC=cl",
    ]

    # 🚫 Descripciones que deben excluirse del resultado
    EXCLUDED_DESC = ["auxiliar", "vigilante privado", "guardia"]

    results = []
    now = datetime.now()

    for base_ou in ALLOWED_OUS:
        try:
            conn.search(base_ou, filter_query, SUBTREE, attributes=attributes)
        except Exception:
            continue

        for entry in conn.entries:
            try:
                sAM = str(entry["sAMAccountName"]) if entry["sAMAccountName"].value else ""
                if not sAM:
                    continue

                # Excluir si descripción coincide con alguno de los términos prohibidos
                desc = str(entry["description"]) if entry["description"].value else ""
                if any(ex in desc.lower() for ex in EXCLUDED_DESC):
                    continue

                # Validar correo (doble seguridad)
                mail = str(entry["mail"]) if entry["mail"].value else ""
                if not mail or "@" not in mail:
                    continue

                display = str(entry["displayName"]) if entry["displayName"].value else sAM
                dept = str(entry["department"]) if entry["department"].value else ""
                expiry_raw = entry["msDS-UserPasswordExpiryTimeComputed"].value
                expiry_dt = msds_to_datetime(expiry_raw)
                if not expiry_dt:
                    continue

                dias_restantes = (expiry_dt - now).days
                results.append({
                    "usuario": sAM,
                    "nombre": display,
                    "correo": mail,
                    "departamento": dept,
                    "dias": dias_restantes,
                    "expira": expiry_dt.strftime("%d/%m/%Y %H:%M"),
                    "descripcion": desc,
                    "dn": str(entry["distinguishedName"]) if entry["distinguishedName"].value else "",
                })
            except Exception:
                continue

    return results




# ==============================
# ENVÍO DE CORREOS (HTML + imagen embebida)
# ==============================
from email.mime.image import MIMEImage

def enviar_correos_lista(usuarios):
    if not usuarios:
        messagebox.showwarning("Aviso", "No hay usuarios seleccionados para enviar correos.")
        return False


def pedir_credenciales_smtp(parent):
    """Ventana modal para pedir correo remitente y contraseña.
    Devuelve (remitente, password) o (None, None) si se cancela.
    Añade opción 'Recordar en esta sesión'."""
    remit_var = tk.StringVar(value=_SMTP_CACHE.get("remitente") or "")
    pass_var = tk.StringVar(value="")
    from_visible_var = tk.StringVar(value=_SMTP_CACHE.get("from_visible") or "")
    remember_var = tk.BooleanVar(value=False)

    # Si hay credenciales fijas configuradas, no pedirlas
    if SMTP_REMITENTE and SMTP_PASSWORD:
        return SMTP_REMITENTE, SMTP_PASSWORD, (_SMTP_CACHE.get("from_visible") or None)

    dlg = tk.Toplevel(parent)
    setup_style(dlg)
    dlg.title("Credenciales SMTP")
    dlg.resizable(False, False)
    dlg.transient(parent)
    dlg.grab_set()
    centrar_ventana(dlg, 520, 260)
    dlg.protocol("WM_DELETE_WINDOW", lambda: (remit_var.set(""), pass_var.set(""), dlg.destroy()))

    frm = ttk.Frame(dlg, padding=12)
    frm.pack(fill="both", expand=True)

    ttk.Label(frm, text="Correo de autenticación (tu cuenta):").grid(row=0, column=0, sticky="w", pady=(4,2))
    e_user = ttk.Entry(frm, textvariable=remit_var, width=44)
    e_user.grid(row=1, column=0, sticky="we")
    e_user.focus()

    ttk.Label(frm, text="Contraseña:").grid(row=2, column=0, sticky="w", pady=(8,2))
    e_pass = ttk.Entry(frm, textvariable=pass_var, show="*", width=44)
    e_pass.grid(row=3, column=0, sticky="we")

    ttk.Label(frm, text="Remitente visible (opcional):").grid(row=4, column=0, sticky="w", pady=(8,2))
    e_from_visible = ttk.Entry(frm, textvariable=from_visible_var, width=44)
    e_from_visible.grid(row=5, column=0, sticky="we")
    ttk.Label(frm, text="Ej.: soporte@capual.cl — si tienes permiso, enviará 'Como'; si no, 'en nombre de'.",
              font=("Segoe UI", 9)).grid(row=6, column=0, sticky="w")

    chk = ttk.Checkbutton(frm, text="Recordar durante esta sesión", variable=remember_var)
    chk.grid(row=7, column=0, sticky="w", pady=(8,4))

    btns = ttk.Frame(frm)
    btns.grid(row=8, column=0, sticky="e", pady=(8,0))

    result = {"ok": False}

    def aceptar():
        user = remit_var.get().strip()
        pw = pass_var.get()
        if not user or "@" not in user:
            messagebox.showwarning("Dato requerido", "Ingresa un correo remitente válido.", parent=dlg)
            return
        if not pw:
            messagebox.showwarning("Dato requerido", "Ingresa la contraseña del remitente.", parent=dlg)
            return
        if remember_var.get():
            _SMTP_CACHE["remitente"] = user
            _SMTP_CACHE["password"] = pw
            _SMTP_CACHE["from_visible"] = from_visible_var.get().strip() or None
        result["ok"] = True
        dlg.destroy()

    def cancelar():
        result["ok"] = False
        dlg.destroy()

    ttk.Button(btns, text="Cancelar", command=cancelar).pack(side="right", padx=(6,0))
    ttk.Button(btns, text="Continuar", command=aceptar).pack(side="right")

    e_pass.bind("<Return>", lambda e: aceptar())
    dlg.bind("<Escape>", lambda e: cancelar())

    dlg.wait_window()
    if result["ok"]:
        return remit_var.get().strip(), pass_var.get(), (from_visible_var.get().strip() or None)
    return None


def enviar_correos_con_progreso(usuarios, parent):
    """Envía correos mostrando una barra de progreso con opción de cancelar.
    Devuelve True si se envió al menos un correo.
    """
    if not usuarios:
        messagebox.showwarning("Aviso", "No hay usuarios seleccionados para enviar correos.", parent=parent)
        return False

    # Elegir método: Outlook (simple) o SMTP
    # Mantener selector si Outlook está disponible; si no, continuar con SMTP directo
    try:
        metodo, enviar_como = seleccionar_metodo_envio(parent)
    except Exception:
        metodo, enviar_como = "smtp", None
    if metodo is None:
        return False
    if metodo == "outlook":
        return enviar_correos_via_outlook(usuarios, parent, enviar_como=enviar_como)

    # Pedir credenciales (From y password); permitir recordar en sesión
    cred = pedir_credenciales_smtp(parent)
    if not cred:
        return False
    if len(cred) == 3:
        remitente, password, from_visible = cred
    else:
        remitente, password = cred
        from_visible = None
    if not remitente or not password:
        return False

    # Confirmación luego de ingresar credenciales
    if not messagebox.askyesno("Confirmar envío", f"¿Desea enviar correos a {len(usuarios)} usuarios?\n\nDesde: {remitente}", parent=parent):
        return False

    progress_win = tk.Toplevel(parent)
    setup_style(progress_win)
    progress_win.title("Enviando correos…")
    progress_win.transient(parent)
    progress_win.grab_set()
    centrar_ventana(progress_win, 420, 140)

    frm = ttk.Frame(progress_win, padding=12)
    frm.pack(fill="both", expand=True)

    lbl = ttk.Label(frm, text="Preparando…")
    lbl.pack(fill="x", pady=(0,8))
    pbar = ttk.Progressbar(frm, mode="determinate", maximum=len(usuarios))
    pbar.pack(fill="x")

    cancel = tk.BooleanVar(value=False)
    def do_cancel():
        cancel.set(True)
    ttk.Button(frm, text="Cancelar", command=do_cancel).pack(pady=(10,0))

    parent.update_idletasks()

    enviados = 0
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(remitente, password)

        for idx, u in enumerate(usuarios, start=1):
            if cancel.get():
                break
            correo = u.get("correo")
            if not correo or "@" not in correo:
                continue

            msg = MIMEMultipart("related")
            # From visible (si se indicó) y Sender para "en nombre de"
            visible = from_visible or remitente
            msg["From"] = visible
            if visible and visible.lower() != remitente.lower():
                msg["Sender"] = remitente
            msg["To"] = correo
            msg["Subject"] = "⚠️ Aviso: Tu contraseña está próxima a expirar"

            html_body = f"""
            <html>
            <body style="font-family:Segoe UI, sans-serif; color:#333;">
                <p>Estimado/a <b>{u.get('nombre','')}</b>,</p>
                <p>Tu contraseña expira en <b>{u.get('dias','-')} días</b> (el {u.get('expira','-')}).<br>
                Por favor, actualízala antes de que caduque para evitar bloqueos de acceso.</p>
                <p><b>Para cambiar tu contraseña:</b><br>
                Presiona <i>Ctrl + Alt + Supr</i> y selecciona la opción "Cambiar contraseña".</p>
                <p style="text-align:center;">
                    <img src="cid:img_teclas" alt="Instrucciones Ctrl+Alt+Supr" width="420">
                </p>
                <p>Si tienes problemas, comunícate con:<br>
                - Eduardo L. (Nexo 4006)<br>
                - Ignacio C. (Nexo 4018)<br>
                Departamento de Servicios TI</p>
                <p>🏢 Departamento: {u.get('departamento','No especificado')}<br>
                👤 Usuario: {u.get('usuario','')}</p>
                <p>Saludos cordiales,<br>
                <b>Departamento de Soporte TI</b><br>
                Capual - Cooperativa de Ahorro y Crédito</p>
            </body>
            </html>
            """
            msg.attach(MIMEText(html_body, "html"))
            try:
                with open(IMG_PATH, "rb") as f:
                    img = MIMEImage(f.read())
                    img.add_header("Content-ID", "<img_teclas>")
                    img.add_header("Content-Disposition", "inline", filename="img_teclas.png")
                    msg.attach(img)
            except FileNotFoundError:
                pass

            try:
                server.send_message(msg)
                enviados += 1
            except Exception:
                pass

            lbl.config(text=f"Enviando {idx}/{len(usuarios)}…")
            pbar['value'] = idx
            progress_win.update_idletasks()

        server.quit()
    except Exception as e:
        messagebox.showerror("Error envío", f"No se pudieron enviar los correos:\n{e}", parent=progress_win)
    finally:
        progress_win.destroy()

    if enviados > 0:
        messagebox.showinfo("Envío finalizado", f"Correos enviados: {enviados}", parent=parent)
        return True
    else:
        return False


# ==============================
# VENTANAS / FLUJO
# ==============================
root_all = None

def ventana_login():
    global root_all
    login_win = tk.Tk()
    root_all = login_win
    login_win.title(f"Iniciar sesión - {APP_CREDITOS}")
    setup_style(login_win)
    centrar_ventana(login_win, 420, 230)
    login_win.resizable(False, False)
    login_win.protocol("WM_DELETE_WINDOW", lambda: confirmar_y_cerrar(login_win))

    frm = ttk.Frame(login_win, padding=12)
    frm.pack(fill="both", expand=True)

    frm.columnconfigure(0, weight=0)
    frm.columnconfigure(1, weight=1)

    ttk.Label(frm, text="Usuario: ").grid(row=0, column=0, sticky="e", pady=(8,4), padx=(4,4))
    usuario_entry = ttk.Entry(frm, width=35)
    usuario_entry.grid(row=0, column=1, pady=(8,4), padx=(4,4))
    usuario_entry.focus()

    ttk.Label(frm, text="Contraseña: ").grid(row=1, column=0, sticky="e", pady=(4,8), padx=(4,4))
    pass_entry = ttk.Entry(frm, width=35, show="*")
    pass_entry.grid(row=1, column=1, pady=(4,8), padx=(4,4))

    status_lbl = ttk.Label(frm, text="")
    status_lbl.grid(row=2, column=0, columnspan=2, pady=(8,4))

    btn_frame = ttk.Frame(frm)
    btn_frame.grid(row=3, column=0, columnspan=2, pady=(10,4))

    def intentar_login():
        user = usuario_entry.get().strip()
        pw = pass_entry.get().strip()
        if not user or not pw:
            messagebox.showwarning("Datos incompletos", "Debes ingresar usuario y contraseña.")
            return
        try:
            status_lbl.config(text="Conectando al AD...")
            login_win.update_idletasks()
            conn = conectar_ldap(user, pw)
            login_win.destroy()
            ventana_principal(conn)
        except Exception:
            messagebox.showerror("Error", "Credenciales inválidas o no se pudo conectar al AD. Intenta nuevamente.")
            usuario_entry.focus()
            status_lbl.config(text="")

    ttk.Button(btn_frame, text="Iniciar sesión", command=intentar_login).grid(row=0, column=0, padx=6)
    ttk.Button(btn_frame, text="Salir", command=lambda: confirmar_y_cerrar(login_win)).grid(row=0, column=1, padx=6)

    login_win.mainloop()

# -----------------------------
# Ventana principal / Menú
# -----------------------------
def ventana_principal(conn):
    global root_all
    main_win = tk.Tk()
    root_all = main_win
    main_win.title(f"Menú Principal - {APP_CREDITOS}")
    setup_style(main_win)
    centrar_ventana(main_win, 560, 300)
    main_win.protocol("WM_DELETE_WINDOW", lambda: confirmar_y_cerrar(main_win))

    frm = ttk.Frame(main_win, padding=20)
    frm.pack(expand=True, fill="both")

    ttk.Label(frm, text="Seleccione una opción", font=("Segoe UI", 12)).pack(pady=(0,12))

    btn1 = ttk.Button(frm, text="📋 Usuarios próximos a expirar", width=36,
                      command=lambda: abrir_usuarios_proximos(main_win, conn))
    btn1.pack(pady=8)

    btn2 = ttk.Button(frm, text="📊 Dashboard de contraseñas (gráfico circular)", width=36,
                      command=lambda: abrir_dashboard(main_win, conn))
    btn2.pack(pady=8)

    # Nuevo botón: búsqueda global por nombre/usuario/correo
    btn3 = ttk.Button(frm, text="🔎 Buscar por nombre o correo", width=36,
                      command=lambda: abrir_busqueda_usuario(main_win, conn))
    btn3.pack(pady=8)

    ttk.Label(frm, text=APP_CREDITOS, font=("Segoe UI", 9)).pack(side="bottom", pady=(12,0))

    main_win.mainloop()

# -----------------------------
# Usuarios próximos a expirar
# (se mantiene funcionalidad previa con orden/filtrado implementados en otra función)
# -----------------------------
def abrir_usuarios_proximos(parent_win, conn):
    parent_win.withdraw()
    win = tk.Toplevel()
    win.title("Usuarios próximos a expirar")
    setup_style(win)
    centrar_ventana(win, 980, 540)
    try:
        win.minsize(900, 520)
    except Exception:
        pass
    win.protocol("WM_DELETE_WINDOW", lambda: on_close_subwindow(win, parent_win))
    frame = ttk.Frame(win, padding=10)
    frame.pack(fill="both", expand=True)

    dias = simpledialog.askinteger("Filtro de días", "¿Cuántos días antes del vencimiento deseas mostrar?", minvalue=1, maxvalue=180, parent=win)
    if not dias:
        dias = 10

    all_users = consultar_usuarios(conn)
    filtered = [u for u in all_users if 0 <= u["dias"] <= dias]

    cols = ("Sel", "Usuario", "Nombre", "Correo", "Departamento", "Días restantes", "Fecha de expiración")
    # Reducimos altura inicial para dejar espacio a los botones siempre visibles
    tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="none", height=15)
    for c in cols:
        tree.heading(c, text=c)
        if c == "Sel":
            tree.column(c, width=60, anchor="center")
        elif c == "Días restantes":
            tree.column(c, width=110, anchor="center")
        else:
            tree.column(c, width=150, anchor="w")
    tree.pack(side="top", fill="both", expand=True)

    seleccion = {}
    item_to_user = {}

    def insertar_datos():
        for row in tree.get_children():
            tree.delete(row)
        seleccion.clear()
        item_to_user.clear()
        for u in filtered:
            vals = ("☐", u["usuario"], u["nombre"], u["correo"], u["departamento"], str(u["dias"]), u["expira"]) 
            iid = tree.insert("", "end", values=vals)
            seleccion[iid] = False
            item_to_user[iid] = u

    insertar_datos()

    # Aplicar comportamiento estándar
    def _ver_props(u):
        ver_propiedades_usuario(win, conn, u)
    make_treeview_standard(tree, cols, item_to_user, _ver_props, seleccion)

    btn_frame = ttk.Frame(frame)
    btn_frame.pack(fill="x", pady=8)

    # Exportadores (CSV/Excel)
    def exportar_csv():
        filas = [tree.item(i, 'values') for i in tree.get_children()]
        if not filas:
            messagebox.showinfo("Exportar", "No hay datos para exportar.", parent=win)
            return
        path = filedialog.asksaveasfilename(parent=win, defaultextension=".csv", filetypes=[("CSV","*.csv")])
        if not path:
            return
        import csv
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(cols)
            writer.writerows(filas)
        messagebox.showinfo("Exportar", "Archivo CSV exportado.", parent=win)

    def exportar_excel():
        export_tree_to_excel(win, tree, titulo=f"Usuarios próximos ({dias} días)")

    def enviar_seleccionados():
        usuarios_sel = [item_to_user[iid] for iid, sel in seleccion.items() if sel]
        if not usuarios_sel:
            messagebox.showwarning("Sin seleccionados", "No hay usuarios seleccionados. Marca con el checkbox en la columna 'Sel'.")
            return
        enviado = enviar_correos_con_progreso(usuarios_sel, win)
        if enviado:
            if messagebox.askyesno("Nueva búsqueda", "¿Deseas hacer una nueva consulta con otro filtro de días?"):
                new_dias = simpledialog.askinteger("Filtro de días", "¿Cuántos días antes del vencimiento deseas mostrar?", minvalue=1, maxvalue=180, parent=win)
                if not new_dias:
                    new_dias = dias
                new_filtered = [u for u in all_users if 0 <= u["dias"] <= new_dias]
                filtered.clear()
                filtered.extend(new_filtered)
                insertar_datos()
            else:
                win.destroy()
                parent_win.deiconify()
        else:
            if messagebox.askyesno("Continuar", "¿Deseas hacer otra búsqueda?"):
                new_dias = simpledialog.askinteger("Filtro de días", "¿Cuántos días antes del vencimiento deseas mostrar?", minvalue=1, maxvalue=180, parent=win)
                if not new_dias:
                    new_dias = dias
                new_filtered = [u for u in all_users if 0 <= u["dias"] <= new_dias]
                filtered.clear()
                filtered.extend(new_filtered)
                insertar_datos()
            else:
                win.destroy()
                parent_win.deiconify()

    def regresar():
        win.destroy()
        parent_win.deiconify()

    # Zona izquierda: exportaciones
    exp_frame = ttk.Frame(btn_frame)
    exp_frame.pack(side="left")
    ttk.Button(exp_frame, text="Exportar Excel", command=exportar_excel).pack(side="left", padx=6)
    ttk.Button(exp_frame, text="Exportar CSV", command=exportar_csv).pack(side="left", padx=6)

    ttk.Button(btn_frame, text="Enviar correo a seleccionados", command=enviar_seleccionados).pack(side="left", padx=8)
    ttk.Button(btn_frame, text="Regresar", command=regresar).pack(side="right", padx=8)

# -----------------------------
# Dashboard mejorado e interactivo
# -----------------------------
def abrir_dashboard(parent_win, conn):
    """Dashboard V2: filtros vivos (estado + rango de días), donut y nuevo histograma,
    y Top 10 urgentes con doble clic a propiedades."""
    parent_win.withdraw()
    win = tk.Toplevel()
    win.title("Dashboard de contraseñas")
    setup_style(win)
    # Dimensión inicial más alta para asegurar visibilidad del pie de página
    centrar_ventana(win, 1000, 760)
    win.protocol("WM_DELETE_WINDOW", lambda: on_close_subwindow(win, parent_win))

    root = ttk.Frame(win, padding=10)
    root.pack(fill="both", expand=True)

    # Datos base
    all_users = consultar_usuarios(conn)

    # Panel superior: filtros y KPIs
    top = ttk.Frame(root)
    top.pack(fill="x", pady=(0,8))

    # Filtros de estado
    estado_vars = {
        "bien": tk.BooleanVar(value=True),
        "proximos": tk.BooleanVar(value=True),
        "expirados": tk.BooleanVar(value=True),
    }
    filtros = ttk.LabelFrame(top, text="Filtros")
    filtros.pack(side="left", fill="x", expand=True)
    ttk.Checkbutton(filtros, text="Bien (16-90)", variable=estado_vars["bien"]).pack(side="left", padx=(6,6))
    ttk.Checkbutton(filtros, text="Próximos (1-15)", variable=estado_vars["proximos"]).pack(side="left", padx=(6,6))
    ttk.Checkbutton(filtros, text="Expirados (≤0)", variable=estado_vars["expirados"]).pack(side="left", padx=(6,6))

    # Rango de días: -30 .. 90
    rango = ttk.LabelFrame(top, text="Rango de días")
    rango.pack(side="right")
    min_var = tk.IntVar(value=-30)
    max_var = tk.IntVar(value=90)
    frm_rng = ttk.Frame(rango)
    frm_rng.pack(padx=6, pady=4)
    ttk.Label(frm_rng, text="Desde").grid(row=0, column=0, padx=(0,6))
    s_min = ttk.Scale(frm_rng, from_=-30, to=90, orient="horizontal", variable=min_var, length=160)
    s_min.grid(row=0, column=1)
    ttk.Label(frm_rng, textvariable=min_var, width=4, anchor="e").grid(row=0, column=2, padx=(6,0))
    ttk.Label(frm_rng, text="Hasta").grid(row=1, column=0, padx=(0,6), pady=(6,0))
    s_max = ttk.Scale(frm_rng, from_=-30, to=90, orient="horizontal", variable=max_var, length=160)
    s_max.grid(row=1, column=1, pady=(6,0))
    ttk.Label(frm_rng, textvariable=max_var, width=4, anchor="e").grid(row=1, column=2, padx=(6,0))

    # KPIs simples
    kpi = ttk.Frame(root)
    kpi.pack(fill="x", pady=(0,6))
    kpi_bien = ttk.Label(kpi, text="🟢 Bien: 0")
    kpi_bien.pack(side="left", padx=8)
    kpi_prox = ttk.Label(kpi, text="🟡 Próximos: 0")
    kpi_prox.pack(side="left", padx=8)
    kpi_exp = ttk.Label(kpi, text="🔴 Expirados: 0")
    kpi_exp.pack(side="left", padx=8)

    # Zona central: gráficos
    charts = ttk.Frame(root)
    charts.pack(fill="both", expand=True)

    # Donut
    fig_donut = Figure(figsize=(5,3.2), dpi=100)
    ax_donut = fig_donut.add_subplot(111)
    canvas_donut = FigureCanvasTkAgg(fig_donut, master=charts)
    canvas_donut.get_tk_widget().pack(side="left", fill="both", expand=True, padx=(0,6))

    # Histograma
    fig_hist = Figure(figsize=(5,3.2), dpi=100)
    ax_hist = fig_hist.add_subplot(111)
    canvas_hist = FigureCanvasTkAgg(fig_hist, master=charts)
    canvas_hist.get_tk_widget().pack(side="left", fill="both", expand=True)

    # Inferior: Top 10 urgentes
    bottom = ttk.Frame(root)
    bottom.pack(fill="both", expand=False, pady=(8,0))
    cols_top = ("Usuario", "Nombre", "Departamento", "Días", "Fecha")
    # Reducimos altura inicial para garantizar espacio al pie
    tree_top = ttk.Treeview(bottom, columns=cols_top, show="headings", height=6)
    for c in cols_top:
        anchor = "center" if c == "Días" else "w"
        width = 70 if c == "Días" else 160
        tree_top.heading(c, text=c)
        tree_top.column(c, width=width, anchor=anchor)
    tree_top.pack(side="left", fill="both", expand=True)
    sb = ttk.Scrollbar(bottom, orient="vertical", command=tree_top.yview)
    tree_top.configure(yscrollcommand=sb.set)
    sb.pack(side="left", fill="y")

    # Doble clic en Top 10 → propiedades
    def on_top_dclick(event):
        iid = tree_top.focus()
        if not iid:
            return
        u = tree_top.item(iid, "values")
        # map values to user by sAMAccountName
        usuario = u[0]
        for obj in all_users:
            if obj.get("usuario") == usuario:
                ver_propiedades_usuario(win, conn, obj)
                break
    tree_top.bind("<Double-1>", on_top_dclick)

    # Utilidades de cálculo
    def _split_estado(lista):
        b = [u for u in lista if 16 <= u.get("dias", 999) <= 90]
        p = [u for u in lista if 1 <= u.get("dias", 999) <= 15]
        e = [u for u in lista if u.get("dias", 999) <= 0]
        return b, p, e

    def _apply_filters():
        mn = min_var.get(); mx = max_var.get()
        if mn > mx:
            mn, mx = mx, mn
        # subset por rango de días
        subset = [u for u in all_users if isinstance(u.get("dias"), int) and mn <= u["dias"] <= mx]
        b, p, e = _split_estado(subset)
        selected = []
        if estado_vars["bien"].get():
            selected += b
        if estado_vars["proximos"].get():
            selected += p
        if estado_vars["expirados"].get():
            selected += e
        return selected, b, p, e

    # Redibujar vistas
    wedges = []
    colors = ["#6aa84f", "#f1c232", "#e06666"]

    def _refresh():
        nonlocal wedges
        data_all, b, p, e = _apply_filters()
        # KPIs
        kpi_bien.configure(text=f"🟢 Bien: {len(b)}")
        kpi_prox.configure(text=f"🟡 Próximos: {len(p)}")
        kpi_exp.configure(text=f"🔴 Expirados: {len(e)}")

        # Donut
        ax_donut.clear()
        counts = [len(b), len(p), len(e)]
        total = sum(counts)
        if total == 0:
            ax_donut.text(0.5, 0.5, "Sin datos con los filtros", ha='center', va='center')
            wedges = []
        else:
            wedges, _, _ = ax_donut.pie(
                counts, labels=None, colors=colors, startangle=90,
                wedgeprops=dict(width=0.5), autopct="%1.1f%%"
            )
            # Habilitar interacción por clic sobre cada porción
            try:
                for w in wedges:
                    w.set_picker(True)
            except Exception:
                pass
            ax_donut.axis('equal')
            labels = [
                f"Bien (16-90): {counts[0]}",
                f"Próximos (1-15): {counts[1]}",
                f"Expirados (<=0): {counts[2]}",
            ]
            ax_donut.legend(wedges, labels, title="Estados", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
        fig_donut.tight_layout()
        canvas_donut.draw()

        # Histograma de días en el subset seleccionado
        ax_hist.clear()
        dias_vals = [u["dias"] for u in data_all if isinstance(u.get("dias"), int)]
        if dias_vals:
            bins = list(range(min_var.get(), max_var.get()+1, 5))
            if len(bins) < 2:
                bins = 10
            ax_hist.hist(dias_vals, bins=bins, color="#98c379", edgecolor="#2e7d32")
            ax_hist.set_title("Distribución de días restantes")
            ax_hist.set_xlabel("Días")
            ax_hist.set_ylabel("Usuarios")
        else:
            ax_hist.text(0.5, 0.5, "Sin datos para histograma", ha='center', va='center')
        fig_hist.tight_layout()
        canvas_hist.draw()

        # Top 10 urgentes
        for iid in tree_top.get_children():
            tree_top.delete(iid)
        urg = sorted([u for u in data_all if u.get("dias") is not None], key=lambda x: x["dias"])[:10]
        for u in urg:
            tree_top.insert("", "end", values=(u.get("usuario"), u.get("nombre"), u.get("departamento"), u.get("dias"), u.get("expira")))

    # Eventos de filtros (debounce suave)
    pending = {"id": None}

    def schedule_refresh(*_):
        if pending["id"]:
            try:
                win.after_cancel(pending["id"])
            except Exception:
                pass
        pending["id"] = win.after(180, _refresh)

    for v in estado_vars.values():
        v.trace_add("write", lambda *a: schedule_refresh())
    s_min.configure(command=lambda v: schedule_refresh())
    s_max.configure(command=lambda v: schedule_refresh())

    # Ventana de categoría basada en filtros actuales
    def abrir_ventana_categoria(nombre_cat, usuarios_cat):
        modal = tk.Toplevel()
        modal.title(f"{nombre_cat} - Usuarios")
        setup_style(modal)
        centrar_ventana(modal, 980, 520)
        modal.transient(win)
        modal.grab_set()

        frm = ttk.Frame(modal, padding=10)
        frm.pack(fill="both", expand=True)

        # Buscador
        buscador_frame = ttk.Frame(frm)
        buscador_frame.pack(fill="x", pady=(0,8))
        ttk.Label(buscador_frame, text="🔍 Buscar:").pack(side="left", padx=(0,6))
        entrada_busq = ttk.Entry(buscador_frame, width=40)
        entrada_busq.pack(side="left", padx=(0,6))

        # Agregamos columna de selección visual "Sel" como primera columna
        cols = ("Sel", "Usuario", "Nombre", "Correo", "Departamento", "Días restantes", "Fecha de expiración")
        # Reducimos la altura para reservar espacio a la botonera inferior
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=15, selectmode="none")
        for c in cols:
            tree.heading(c, text=c)
            if c == "Sel":
                tree.column(c, width=60, anchor="center")
            elif c == "Días restantes":
                tree.column(c, width=120, anchor="center")
            else:
                tree.column(c, width=140, anchor="w")
        tree.pack(fill="both", expand=True)

        seleccion = {}
        item_to_user = {}

        # Insertar datos
        def insertar_tabla(datos):
            for r in tree.get_children():
                tree.delete(r)
            seleccion.clear()
            item_to_user.clear()
            for u in datos:
                iid = tree.insert("", "end", values=("☐", u["usuario"], u["nombre"], u["correo"],
                                                     u["departamento"], u.get("dias","-"), u.get("expira","-")))
                seleccion[iid] = False
                item_to_user[iid] = u

        insertar_tabla(usuarios_cat)

        # Comportamiento estándar: ordenar, select-all, resalte, doble clic abre propiedades
        def _ver_props(u):
            ver_propiedades_usuario(modal, conn, u)
        make_treeview_standard(tree, cols, item_to_user, _ver_props, seleccion)

        # Filtro
        def filtrar():
            texto = entrada_busq.get().strip().lower()
            for item in tree.get_children():
                valores = tree.item(item, "values")
                visible = any(texto in str(v).lower() for v in valores)
                tree.detach(item) if not visible else tree.reattach(item, '', 'end')

        def limpiar():
            entrada_busq.delete(0, tk.END)
            insertar_tabla(usuarios_cat)

        ttk.Button(buscador_frame, text="Buscar", command=filtrar).pack(side="left", padx=(0,6))
        ttk.Button(buscador_frame, text="Limpiar", command=limpiar).pack(side="left")
        # accesos rápidos
        entrada_busq.bind("<Return>", lambda e: filtrar())
        modal.bind("<Escape>", lambda e: (entrada_busq.delete(0, tk.END), insertar_tabla(usuarios_cat)))

        # Botones inferiores
        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=8)
        ttk.Button(btns, text="Regresar", command=lambda: modal.destroy()).pack(side="right", padx=6)

        # Exportar a Excel la vista actual
        def exportar_excel_local():
            export_tree_to_excel(modal, tree, titulo=f"{nombre_cat} - Usuarios")
        ttk.Button(btns, text="Exportar Excel", command=exportar_excel_local).pack(side="left", padx=6)

        # Solo para “Próximos” y “Expirados”
        if nombre_cat in ("Próximos", "Expirados"):
            def enviar_todos():
                visibles = []
                for item in tree.get_children():
                    usuario_key = tree.item(item, "values")[1]  # columna Usuario (después de Sel)
                    for u in usuarios_cat:
                        if u["usuario"] == usuario_key:
                            visibles.append(u)
                            break
                if not visibles:
                    messagebox.showwarning("Sin destinatarios", "No hay usuarios visibles para enviar.")
                    return
                if messagebox.askyesno("Confirmar envío", f"¿Enviar correos a {len(visibles)} usuarios visibles?"):
                    enviar_correos_con_progreso(visibles, modal)

            def enviar_seleccionados():
                seleccionados = [item_to_user[i] for i, sel in seleccion.items() if sel]
                if not seleccionados:
                    messagebox.showwarning("Sin seleccionados", "No has seleccionado ningún usuario. Marca con el checkbox en 'Sel'.")
                    return
                if messagebox.askyesno("Confirmar envío", f"¿Enviar correos a {len(seleccionados)} usuarios seleccionados?"):
                    enviar_correos_con_progreso(seleccionados, modal)

            ttk.Button(btns, text="Enviar correos a todos", command=enviar_todos).pack(side="left", padx=6)
            ttk.Button(btns, text="Enviar correos seleccionados", command=enviar_seleccionados).pack(side="left", padx=6)



    # Conexión de picking y tooltips para el donut
    tooltip = tk.Label(root, text="", background="#ffffe0", relief="solid", bd=1)
    tooltip.place_forget()

    def on_pick(event):
        if not wedges:
            return
        try:
            idx = wedges.index(event.artist)
        except Exception:
            return
        # Construir subset según filtros actuales y abrir
        _, b, p, e = _apply_filters()
        if idx == 0:
            abrir_ventana_categoria("Bien", b)
        elif idx == 1:
            abrir_ventana_categoria("Próximos", p)
        elif idx == 2:
            abrir_ventana_categoria("Expirados", e)

    def on_move(event):
        if not wedges or not event.inaxes:
            tooltip.place_forget()
            return
        # recrear labels con conteos actuales
        _, b, p, e = _apply_filters()
        labels = [
            f"Bien (16-90): {len(b)}",
            f"Próximos (1-15): {len(p)}",
            f"Expirados (<=0): {len(e)}",
        ]
        found = False
        for i, w in enumerate(wedges):
            contains, _ = w.contains(event)
            if contains:
                tooltip.config(text=labels[i])
                tooltip.place(x=event.x + 20, y=event.y + 20)
                found = True
                break
        if not found:
            tooltip.place_forget()

    canvas_donut.mpl_connect('pick_event', on_pick)
    canvas_donut.mpl_connect('motion_notify_event', on_move)

    # botones: anclados al fondo y acción rápida para abrir la vista filtrada
    def abrir_vista_filtrada():
        data_all, _, _, _ = _apply_filters()
        abrir_ventana_categoria("Vista filtrada", data_all)

    btn_frame = ttk.Frame(root)
    btn_frame.pack(side="bottom", fill="x", pady=8)
    ttk.Button(btn_frame, text="Abrir vista filtrada", command=abrir_vista_filtrada).pack(side="left", padx=8)
    ttk.Button(btn_frame, text="Regresar", command=lambda: (win.destroy(), parent_win.deiconify())).pack(side="right", padx=8)

    # Ajuste dinámico tras construir la UI: calculamos el tamaño requerido
    try:
        win.update_idletasks()
        screen_w = win.winfo_screenwidth()
        screen_h = win.winfo_screenheight()
        req_w = win.winfo_reqwidth()
        req_h = win.winfo_reqheight()
        # Objetivo: al menos 720px alto o lo que requiera el contenido, sin exceder la pantalla
        target_w = min(max(1000, req_w), screen_w - 60)
        target_h = min(max(720, req_h), screen_h - 80)
        centrar_ventana(win, target_w, target_h)
        # Evitar que el usuario reduzca tanto que oculte la botonera
        min_w = min(max(900, int(target_w*0.9)), screen_w - 100)
        min_h = min(max(640, int(target_h*0.85)), screen_h - 120)
        win.minsize(min_w, min_h)
    except Exception:
        pass

    # Inicializar
    _refresh()

# -----------------------------
# Cierre modal
# -----------------------------
def on_close_subwindow(win, parent_win):
    if messagebox.askyesno("Confirmar", "¿Deseas volver al menú principal?"):
        try:
            win.destroy()
        finally:
            parent_win.deiconify()
    else:
        return

# -----------------------------
# Utilidad para escapar filtros LDAP
# -----------------------------
def escape_ldap_filter_value(value: str) -> str:
    """Escapa caracteres especiales para filtros LDAP (RFC4515)."""
    repl = {
        "\\": r"\\5c",
        "*": r"\\2a",
        "(": r"\\28",
        ")": r"\\29",
        "\x00": r"\\00",
    }
    out = []
    for ch in value:
        out.append(repl.get(ch, ch))
    return "".join(out)

# -----------------------------
# Búsqueda global por nombre/usuario/correo en todo el AD
# -----------------------------
def buscar_usuarios_global(conn, termino: str, *, incluir_deshabilitados: bool = False, incluir_pwd_never_expires: bool = True, base_dn: str = None):
    termino = (termino or "").strip()
    if not termino:
        return []

    term = escape_ldap_filter_value(termino)

    # Construir filtro de forma flexible
    filtros = [
        "(objectCategory=person)",
        "(objectClass=user)",
    ]
    if not incluir_deshabilitados:
        filtros.append("(!(userAccountControl:1.2.840.113556.1.4.803:=2))")
    if not incluir_pwd_never_expires:
        filtros.append("(!(userAccountControl:1.2.840.113556.1.4.803:=65536))")

    filtros.append(
        "(|" +
        f"(displayName=*{term}*)" +
        f"(sAMAccountName=*{term}*)" +
        f"(mail=*{term}*)" +
        ")"
    )

    filter_query = "(&" + "".join(filtros) + ")"

    attributes = [
        "sAMAccountName",
        "displayName",
        "mail",
        "msDS-UserPasswordExpiryTimeComputed",
        "department",
        "description",
        "distinguishedName",
    ]

    EXCLUDED_DESC = ["auxiliar", "vigilante privado", "guardia"]

    results = []
    now = datetime.now()
    try:
        conn.search(base_dn or BASE_DN, filter_query, SUBTREE, attributes=attributes)
    except Exception:
        return []

    for entry in conn.entries:
        try:
            sAM = str(entry["sAMAccountName"]) if entry["sAMAccountName"].value else ""
            if not sAM:
                continue

            desc = str(entry["description"]) if entry["description"].value else ""
            # Excluir descripciones específicas; incluir vacías
            if desc and any(ex in desc.lower() for ex in EXCLUDED_DESC):
                continue

            mail = str(entry["mail"]) if entry["mail"].value else ""
            display = str(entry["displayName"]) if entry["displayName"].value else sAM
            dept = str(entry["department"]) if entry["department"].value else ""
            dn = str(entry["distinguishedName"]) if entry["distinguishedName"].value else ""
            expiry_raw = entry["msDS-UserPasswordExpiryTimeComputed"].value
            expiry_dt = msds_to_datetime(expiry_raw)
            dias_restantes = (expiry_dt - now).days if expiry_dt else None

            results.append({
                "usuario": sAM,
                "nombre": display,
                "correo": mail,
                "departamento": dept,
                "dias": dias_restantes if dias_restantes is not None else "-",
                "expira": expiry_dt.strftime("%d/%m/%Y %H:%M") if expiry_dt else "-",
                "descripcion": desc,
                "dn": dn,
            })
        except Exception:
            continue

    return results

# -----------------------------
# Modal: Propiedades del usuario (con foto)
# -----------------------------
def ver_propiedades_usuario(parent_win, conn, usuario_dict: dict):
    dn = (usuario_dict or {}).get("dn") or ""
    sam = (usuario_dict or {}).get("usuario") or ""
    # Atributos a recuperar
    attributes = [
        "displayName",
        "sAMAccountName",
        "mail",
        "department",
        "title",
        "description",
        "userAccountControl",
        "whenCreated",
        "whenChanged",
        "lastLogonTimestamp",
        "pwdLastSet",
        "telephoneNumber",
        "ipPhone",
        "mobile",
        "physicalDeliveryOfficeName",
        "thumbnailPhoto",
        "distinguishedName",
    ]

    entry = None
    try:
        base = dn if dn else BASE_DN
        filt = "(objectClass=person)" if dn else f"(sAMAccountName={sam})"
        conn.search(base, filt, SUBTREE, attributes=attributes)
        if conn.entries:
            entry = conn.entries[0]
    except Exception:
        entry = None

    # Extraer valores seguros
    def _get(attr, default=""):
        try:
            v = entry[attr].value
            if v is None:
                return default
            return str(v)
        except Exception:
            return default

    display = fix_text_encoding(_get("displayName", usuario_dict.get("nombre", sam)))
    sAM = fix_text_encoding(_get("sAMAccountName", sam))
    mail = fix_text_encoding(_get("mail", usuario_dict.get("correo", "")))
    dept = fix_text_encoding(_get("department", usuario_dict.get("departamento", "")))
    title = fix_text_encoding(_get("title", ""))
    desc = fix_text_encoding(_get("description", usuario_dict.get("descripcion", "")))
    tel = fix_text_encoding(_get("telephoneNumber", ""))
    ipphone = fix_text_encoding(_get("ipPhone", ""))
    mobile = fix_text_encoding(_get("mobile", ""))
    office = fix_text_encoding(_get("physicalDeliveryOfficeName", ""))
    dn_full = _get("distinguishedName", dn)  # ya no se muestra
    dias = usuario_dict.get("dias", "-")

    # Convertir timestamps si disponibles
    def _to_dt(v):
        try:
            return msds_to_datetime(v)
        except Exception:
            return None
    last_logon_dt = _to_dt(getattr(entry["lastLogonTimestamp"], 'value', None) if entry else None)
    pwd_last_set_dt = _to_dt(getattr(entry["pwdLastSet"], 'value', None) if entry else None)
    # Fecha de ingreso: usar whenCreated del objeto (aproxima fecha de alta)
    ingreso_raw = None
    try:
        ingreso_raw = entry["whenCreated"].value if entry else None
    except Exception:
        ingreso_raw = None
    ingreso_dt = None
    if ingreso_raw:
        if isinstance(ingreso_raw, datetime):
            ingreso_dt = ingreso_raw
        else:
            try:
                # ldap3 puede devolver string ISO
                ingreso_dt = datetime.fromisoformat(str(ingreso_raw).replace('Z','+00:00'))
            except Exception:
                try:
                    ingreso_dt = datetime.strptime(str(ingreso_raw), "%Y%m%d%H%M%S.%fZ")
                except Exception:
                    ingreso_dt = None
    # Mostrar solo fecha (sin hora)
    ingreso_str = ingreso_dt.strftime("%d/%m/%Y") if ingreso_dt else "-"

    # Foto
    photo_data = None
    try:
        if entry and entry["thumbnailPhoto"].value:
            photo_data = entry["thumbnailPhoto"].value
    except Exception:
        photo_data = None

    # Construir modal
    dlg = tk.Toplevel(parent_win)
    setup_style(dlg)
    dlg.title(f"Información de: {display}")
    dlg.transient(parent_win)
    dlg.grab_set()
    # Altura base más generosa para evitar que la botonera quede fuera del marco
    centrar_ventana(dlg, 740, 520)

    frm = ttk.Frame(dlg, padding=12)
    frm.pack(fill="both", expand=True)

    # Top: Encabezado llamativo con fondo
    header = tk.Frame(frm, bg="#e6f3ec")
    header.pack(fill="x", pady=(0,10))
    photo_label = tk.Label(header, bg="#e6f3ec")
    photo_label.pack(side="left", padx=(8,12), pady=6)
    title_box = tk.Frame(header, bg="#e6f3ec")
    title_box.pack(side="left", pady=6)
    name_lbl = tk.Label(title_box, text=display, font=("Segoe UI Semibold", 16), bg="#e6f3ec", fg="#20302a")
    name_lbl.pack(anchor="w")
    if title:
        sub_lbl = tk.Label(title_box, text=title, font=("Segoe UI", 11), bg="#e6f3ec", fg="#415a4f")
        sub_lbl.pack(anchor="w")

    # Render foto si existe
    try:
        if photo_data:
            from io import BytesIO
            from PIL import Image, ImageTk  # type: ignore
            im = Image.open(BytesIO(photo_data))
            im = im.convert("RGB")
            im.thumbnail((128,128), Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS)
            ph = ImageTk.PhotoImage(im)
            photo_label.configure(image=ph)
            dlg._user_photo_ref = ph  # evitar GC
        else:
            photo_label.configure(text="(sin foto)", fg="#6b6b6b")
    except Exception:
        photo_label.configure(text="(sin foto)", fg="#6b6b6b")

    # Grid de propiedades
    grid = ttk.Frame(frm)
    grid.pack(fill="both", expand=True)
    try:
        grid.columnconfigure(0, weight=0)
        grid.columnconfigure(1, weight=1)
    except Exception:
        pass

    def add_row(r, label, value, *, link_mail=False):
        ttk.Label(grid, text=label+":", font=("Segoe UI", 10, "bold")).grid(row=r, column=0, sticky="e", padx=(0,6), pady=2)
        if link_mail and value:
            # Crear enlace clicable sin forzar 'bg' (ttk usa temas y puede no devolver color de fondo)
            try:
                lbl = tk.Label(grid, text=value, fg="#1a73e8", cursor="hand2")
                def _open_mail(evt=None, m=value):
                    try:
                        os.startfile(f"mailto:{m}")
                    except Exception:
                        pass
                lbl.bind("<Button-1>", _open_mail)
                lbl.grid(row=r, column=1, sticky="w", pady=2)
            except Exception:
                ttk.Label(grid, text=str(value)).grid(row=r, column=1, sticky="w", pady=2)
        else:
            ttk.Label(grid, text=("" if value is None else str(value))).grid(row=r, column=1, sticky="w", pady=2)

    row = 0
    try:
        add_row(row, "Usuario", sAM); row += 1
        add_row(row, "Correo", mail, link_mail=True); row += 1
        add_row(row, "Departamento", dept); row += 1
        if title:
            add_row(row, "Cargo", title); row += 1
        if tel:
            add_row(row, "Teléfono", tel); row += 1
        # Nexo / IP: ipPhone si existe; si no, intentar extraer extensión de Teléfono
        nexo = ipphone
        if not nexo:
            try:
                m = re.search(r"\b(\d{4})\b", str(tel or ""))
                nexo = m.group(1) if m else ""
            except Exception:
                nexo = ""
        if nexo:
            add_row(row, "Nexo/IP", nexo); row += 1
        if mobile:
            add_row(row, "Móvil", mobile); row += 1
        if office:
            add_row(row, "Oficina", office); row += 1
        if desc:
            add_row(row, "Descripción", desc); row += 1
        add_row(row, "Días restantes", str(dias)); row += 1
        add_row(row, "Ingreso", ingreso_str); row += 1
        if last_logon_dt:
            add_row(row, "Último inicio sesión", last_logon_dt.strftime("%d/%m/%Y %H:%M")); row += 1
        if pwd_last_set_dt:
            add_row(row, "Pwd actualizada", pwd_last_set_dt.strftime("%d/%m/%Y %H:%M")); row += 1
    except Exception:
        # Ante cualquier imprevisto, no romper el modal
        pass

    # Botones
    btns = ttk.Frame(frm)
    btns.pack(fill="x", pady=(10,0))

    def _copy(text):
        try:
            dlg.clipboard_clear(); dlg.clipboard_append(text)
        except Exception:
            pass

    # Solo dejamos 'Copiar usuario' como pediste
    ttk.Button(btns, text="Copiar usuario", command=lambda: _copy(sAM)).pack(side="left", padx=4)
    ttk.Button(btns, text="Cerrar", command=lambda: dlg.destroy()).pack(side="right")

    dlg.bind("<Escape>", lambda e: dlg.destroy())

    # Ajustar tamaño final según contenido para no ocultar la botonera; limitar a alto de pantalla
    try:
        dlg.update_idletasks()
        req_w = dlg.winfo_reqwidth()
        req_h = dlg.winfo_reqheight()
        scr_w = dlg.winfo_screenwidth()
        scr_h = dlg.winfo_screenheight()
        final_w = max(740, min(req_w + 20, scr_w - 80))
        final_h = max(520, min(req_h + 20, scr_h - 120))
        centrar_ventana(dlg, final_w, final_h)
        # Fijar un mínimo razonable por si se redimensiona accidentalmente
        dlg.minsize(680, 460)
    except Exception:
        pass

# -----------------------------
# Ventana de búsqueda global
# -----------------------------
def abrir_busqueda_usuario(parent_win, conn):
    parent_win.withdraw()
    win = tk.Toplevel()
    win.title("Buscar usuario por nombre, usuario o correo")
    setup_style(win)
    centrar_ventana(win, 980, 560)
    win.protocol("WM_DELETE_WINDOW", lambda: on_close_subwindow(win, parent_win))

    frame = ttk.Frame(win, padding=10)
    frame.pack(fill="both", expand=True)

    # Barra de búsqueda
    top = ttk.Frame(frame)
    top.pack(fill="x", pady=(0,8))
    ttk.Label(top, text="🔎 Término a buscar:").pack(side="left", padx=(0,6))
    entrada = ttk.Entry(top, width=34)
    entrada.pack(side="left")

    # Búsqueda simplificada (solo término)
    def realizar_busqueda():
        termino = entrada.get().strip()
        if not termino:
            messagebox.showwarning("Dato requerido", "Ingresa un nombre, usuario o correo a buscar.")
            return
        start = datetime.now()
        datos = buscar_usuarios_global(conn, termino)
        insertar_datos(datos)
        elapsed = (datetime.now() - start).total_seconds()
        status_lbl.config(text=f"Coincidencias: {len(datos)}   •   Tiempo: {elapsed:.2f}s")

    ttk.Button(top, text="Buscar", command=realizar_busqueda).pack(side="left", padx=8)
    ttk.Button(top, text="Limpiar", command=lambda: (entrada.delete(0, tk.END), insertar_datos([]), status_lbl.config(text=""))).pack(side="left")

    status_lbl = ttk.Label(frame, text="")
    status_lbl.pack(fill="x", pady=(0,6))

    cols = ("Sel", "Usuario", "Nombre", "Correo", "Departamento", "Días restantes", "Fecha de expiración")
    # Reducimos altura para garantizar visibilidad de la botonera inferior
    tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="none", height=15)
    for c in cols:
        tree.heading(c, text=c)
        if c == "Sel":
            tree.column(c, width=60, anchor="center")
        elif c in ("Días restantes", "Fecha de expiración"):
            tree.column(c, width=130, anchor="center")
        else:
            tree.column(c, width=160, anchor="w")
    tree.pack(side="top", fill="both", expand=True)

    seleccion = {}
    item_to_user = {}

    def insertar_datos(datos):
        for r in tree.get_children():
            tree.delete(r)
        seleccion.clear()
        item_to_user.clear()
        for u in datos:
            vals = ("☐", u["usuario"], u["nombre"], u["correo"], u["departamento"], str(u["dias"]), u["expira"]) 
            iid = tree.insert("", "end", values=vals)
            seleccion[iid] = False
            item_to_user[iid] = u
    # altura fija: sin auto-resize

    # Comportamiento estándar: ordenar, seleccionar con check y doble clic abre propiedades
    def _ver_props(u):
        ver_propiedades_usuario(win, conn, u)
    make_treeview_standard(tree, cols, item_to_user, _ver_props, seleccion)

    # Botonera
    btns = ttk.Frame(frame)
    btns.pack(fill="x", pady=8)

    def enviar_sel():
        usuarios_sel = [item_to_user[iid] for iid, sel in seleccion.items() if sel]
        if not usuarios_sel:
            messagebox.showwarning("Sin seleccionados", "No hay usuarios seleccionados. Marca con el checkbox en 'Sel'.")
            return
        enviar_correos_con_progreso(usuarios_sel, win)

    def export_csv():
        rows = [tree.item(i, 'values') for i in tree.get_children()]
        if not rows:
            messagebox.showinfo("Exportar", "No hay datos para exportar.", parent=win)
            return
        path = filedialog.asksaveasfilename(parent=win, defaultextension=".csv", filetypes=[("CSV","*.csv")])
        if not path:
            return
        import csv
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(cols)
            writer.writerows(rows)
        messagebox.showinfo("Exportar", "Archivo CSV exportado.", parent=win)

    def copiar_portapapeles():
        rows = [tree.item(i, 'values') for i in tree.get_children()]
        if not rows:
            return
        text = "\n".join("\t".join(map(str, r)) for r in rows)
        try:
            win.clipboard_clear()
            win.clipboard_append(text)
        except Exception:
            pass

    ttk.Button(btns, text="Exportar CSV", command=export_csv).pack(side="left", padx=6)
    ttk.Button(btns, text="Exportar Excel", command=lambda: export_tree_to_excel(win, tree, titulo="Búsqueda de usuarios")).pack(side="left", padx=6)
    ttk.Button(btns, text="Copiar", command=copiar_portapapeles).pack(side="left", padx=6)
    ttk.Button(btns, text="Enviar correo a seleccionados", command=enviar_sel).pack(side="left", padx=6)
    ttk.Button(btns, text="Regresar", command=lambda: (win.destroy(), parent_win.deiconify())).pack(side="right", padx=6)

    # Atajos: Enter para buscar, Escape para limpiar
    entrada.bind("<Return>", lambda e: realizar_busqueda())
    win.bind("<Escape>", lambda e: (entrada.delete(0, tk.END), insertar_datos([]), status_lbl.config(text="")))

# ==============================
# PUNTO DE ENTRADA
# ==============================
if __name__ == "__main__":
    ventana_login()
